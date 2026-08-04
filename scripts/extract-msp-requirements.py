#!/usr/bin/env python3
"""
Extract MSP requirements from the official Excel self-assessment
"""

import json
from openpyxl import load_workbook

def extract_requirements(xlsx_path):
    """Extract all requirements from MSP Technical Validation tab"""
    wb = load_workbook(xlsx_path)

    # Find the Technical Validation sheet
    sheet = None
    for name in wb.sheetnames:
        if 'Technical Validation' in name or 'MSP Technical' in name:
            sheet = wb[name]
            break

    if not sheet:
        print("Available sheets:", wb.sheetnames)
        raise ValueError("Could not find MSP Technical Validation sheet")

    requirements = []

    # Iterate through rows to find requirements
    # Format is typically: ID | Name | Description | Evidence | etc.
    header_row = None
    for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        row_text = ' '.join(str(cell or '') for cell in row).lower()
        if 'requirement' in row_text and ('id' in row_text or 'name' in row_text):
            header_row = idx
            print(f"Found header row at line {idx}")
            print("Headers:", [cell for cell in row if cell])
            break

    if not header_row:
        print("Could not find header row, dumping first 20 rows:")
        for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True), 1):
            print(f"Row {idx}:", [cell for cell in row if cell])
        raise ValueError("Could not find requirement header row")

    # Parse requirements starting after header
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        # Skip empty rows
        if not any(row):
            continue

        # Get ID from first column (typically)
        req_id = str(row[0] or '').strip()

        # Skip if not a valid ID format
        if not req_id or not any(c.isalpha() for c in req_id):
            continue

        # Extract data (adjust column indices based on actual spreadsheet)
        requirement = {
            'id': req_id,
            'name': str(row[1] or '').strip() if len(row) > 1 else '',
            'description': str(row[2] or '').strip() if len(row) > 2 else '',
            'category': req_id.split('-')[0] if '-' in req_id else 'UNKNOWN',
        }

        # Only add if has ID and name
        if requirement['id'] and requirement['name']:
            requirements.append(requirement)

    return requirements

if __name__ == '__main__':
    import sys

    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else \
        '~/Downloads/AWS Managed Service Provider (MSP) Program Self-Assessment.xlsx'

    xlsx_path = xlsx_path.replace('~', '/Users/tim')

    try:
        reqs = extract_requirements(xlsx_path)
        print(f"\nFound {len(reqs)} requirements:")

        # Group by category
        by_category = {}
        for req in reqs:
            cat = req['category']
            if cat not in by_category:
                by_category[cat] = []
            by_category[cat].append(req)

        for cat in sorted(by_category.keys()):
            print(f"\n{cat} ({len(by_category[cat])}):")
            for req in by_category[cat]:
                print(f"  {req['id']}: {req['name'][:60]}")

        # Save to JSON
        output_path = '/Users/tim/repos/flexion-msp-readiness/msp-requirements-extracted.json'
        with open(output_path, 'w') as f:
            json.dump(reqs, f, indent=2)
        print(f"\nSaved to: {output_path}")

    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
